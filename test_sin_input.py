import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from datetime import datetime
import os
import time
import random
import threading

class TestBalanzaSinInput:
    def __init__(self):
        self.data_list = []
        self.lote_number = ""
        self.root = None
        self.control_frame = None
        self.status_label = None
        self.data_text = None
        self.contador = 0
        self.running = False
        self.paused = False
        
    def solicitar_numero_lote(self):
        """
        Solicita el número de lote al usuario mediante interfaz gráfica
        """
        root = tk.Tk()
        root.title("Configuración de Lote - Prueba")
        root.geometry("400x200")
        root.eval('tk::PlaceWindow . center')
        
        frame = tk.Frame(root, padx=20, pady=20)
        frame.pack(expand=True, fill='both')
        
        tk.Label(frame, text="SISTEMA DE ADQUISICIÓN DE DATOS (PRUEBA)", 
                font=("Arial", 16, "bold")).pack(pady=10)
        
        tk.Label(frame, text="Ingrese el número de lote:", 
                font=("Arial", 12)).pack(pady=10)
        
        lote_entry = tk.Entry(frame, font=("Arial", 12), width=20)
        lote_entry.pack(pady=10)
        lote_entry.focus()
        
        def on_ok():
            self.lote_number = lote_entry.get()
            if self.lote_number:
                root.destroy()
            else:
                messagebox.showwarning("Advertencia", "Por favor ingrese un número de lote")
        
        def on_cancel():
            self.lote_number = None
            root.destroy()
        
        button_frame = tk.Frame(frame)
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="Aceptar", command=on_ok, 
                font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Cancelar", command=on_cancel, 
                font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)
        
        root.bind('<Return>', lambda e: on_ok())
        root.mainloop()
        return self.lote_number
    
    def crear_interfaz_principal(self):
        """
        Crea la interfaz gráfica principal del sistema de prueba
        """
        self.root = tk.Tk()
        self.root.title("Sistema de Adquisición de Datos de Balanza - PRUEBA")
        self.root.geometry("800x600")
        self.root.eval('tk::PlaceWindow . center')
        
        # Frame principal
        main_frame = tk.Frame(self.root, padx=10, pady=10)
        main_frame.pack(expand=True, fill='both')
        
        # Título
        title_label = tk.Label(main_frame, text="SISTEMA DE ADQUISICIÓN DE DATOS DE BALANZA (PRUEBA)", 
                              font=("Arial", 18, "bold"), fg="darkblue")
        title_label.pack(pady=10)
        
        # Información del lote
        lote_frame = tk.Frame(main_frame)
        lote_frame.pack(fill='x', pady=5)
        
        tk.Label(lote_frame, text=f"Lote Actual: {self.lote_number}", 
                font=("Arial", 12, "bold"), fg="green").pack(side=tk.LEFT)
        
        tk.Label(lote_frame, text="MODO PRUEBA - Datos Simulados", 
                font=("Arial", 10), fg="orange").pack(side=tk.RIGHT)
        
        # Área de datos
        data_frame = tk.Frame(main_frame)
        data_frame.pack(expand=True, fill='both', pady=10)
        
        tk.Label(data_frame, text="Datos Simulados de la Balanza:", 
                font=("Arial", 12, "bold")).pack(anchor='w')
        
        # Text widget con scrollbar
        text_frame = tk.Frame(data_frame)
        text_frame.pack(expand=True, fill='both', pady=5)
        
        self.data_text = tk.Text(text_frame, height=15, font=("Consolas", 10))
        scrollbar = tk.Scrollbar(text_frame, orient="vertical", command=self.data_text.yview)
        self.data_text.configure(yscrollcommand=scrollbar.set)
        
        self.data_text.pack(side=tk.LEFT, expand=True, fill='both')
        scrollbar.pack(side=tk.RIGHT, fill='y')
        
        # Frame de control
        self.control_frame = tk.Frame(main_frame)
        self.control_frame.pack(fill='x', pady=10)
        
        # Botones de control
        button_frame = tk.Frame(self.control_frame)
        button_frame.pack()
        
        self.btn_iniciar = tk.Button(button_frame, text="INICIAR SIMULACIÓN", 
                                   command=self.iniciar_simulacion, 
                                   font=("Arial", 12, "bold"), 
                                   bg="green", fg="white", width=18)
        self.btn_iniciar.pack(side=tk.LEFT, padx=5)
        
        self.btn_detener = tk.Button(button_frame, text="DETENER Y FINALIZAR", 
                                   command=self.detener_y_finalizar, 
                                   font=("Arial", 12, "bold"), 
                                   bg="red", fg="white", width=20, state='disabled')
        self.btn_detener.pack(side=tk.LEFT, padx=5)
        
        self.btn_pausar = tk.Button(button_frame, text="PAUSAR", 
                                  command=self.pausar_reanudar, 
                                  font=("Arial", 12, "bold"), 
                                  bg="orange", fg="white", width=15, state='disabled')
        self.btn_pausar.pack(side=tk.LEFT, padx=5)
        
        # Status bar
        self.status_label = tk.Label(main_frame, text="Sistema de prueba listo. Presione 'INICIAR SIMULACIÓN' para comenzar.", 
                                   font=("Arial", 10), fg="blue", relief=tk.SUNKEN, anchor='w')
        self.status_label.pack(fill='x', pady=5)
        
        # Configurar cierre de ventana
        self.root.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion)
    
    def actualizar_status(self, mensaje):
        """
        Actualiza el mensaje de estado
        """
        if self.status_label:
            self.status_label.config(text=mensaje)
            self.root.update()
    
    def agregar_dato(self, datos):
        """
        Agrega un dato al área de texto
        """
        if self.data_text:
            self.contador += 1
            timestamp = datetime.now().strftime('%H:%M:%S')
            texto = f"[{timestamp}] Lectura #{self.contador}: Peso={datos['peso_bruto']} kg, Tara={datos['tara']} kg, Neto={datos['peso_neto']} kg\n"
            self.data_text.insert(tk.END, texto)
            self.data_text.see(tk.END)
            self.root.update()
    
    def iniciar_simulacion(self):
        """
        Inicia la simulación de datos en un hilo separado
        """
        self.running = True
        self.paused = False
        self.btn_iniciar.config(state='disabled')
        self.btn_detener.config(state='normal')
        self.btn_pausar.config(state='normal')
        
        # Iniciar hilo de simulación
        self.thread_simulacion = threading.Thread(target=self.hilo_simulacion, daemon=True)
        self.thread_simulacion.start()
        
        self.actualizar_status("Simulando datos de la balanza... Presione 'DETENER' para parar.")
    
    def hilo_simulacion(self):
        """
        Hilo principal de simulación de datos
        """
        while self.running:
            try:
                if not self.paused:
                    datos = self.simular_datos()
                    if datos:
                        self.agregar_dato(datos)
                time.sleep(1)
            except Exception as e:
                self.actualizar_status(f"Error en simulación: {e}")
                break
    
    def detener_y_finalizar(self):
        """
        Detiene la simulación y finaliza el programa
        """
        self.running = False
        self.actualizar_status("Deteniendo simulación...")
        
        # Guardar datos
        if self.data_list:
            self.actualizar_status("Guardando datos...")
            if self.guardar_datos_excel():
                self.actualizar_status("Datos guardados exitosamente. Cerrando aplicación...")
            else:
                self.actualizar_status("Error al guardar datos. Cerrando aplicación...")
        else:
            self.actualizar_status("No hay datos para guardar. Cerrando aplicación...")
        
        # Cerrar aplicación después de un breve delay
        self.root.after(2000, self.root.destroy)
    
    def pausar_reanudar(self):
        """
        Pausa o reanuda la simulación
        """
        if self.paused:
            # Reanudar
            self.paused = False
            self.btn_pausar.config(text="PAUSAR")
            self.actualizar_status("Simulación reanudada. Generando datos...")
        else:
            # Pausar
            self.paused = True
            self.btn_pausar.config(text="REANUDAR")
            self.actualizar_status("Simulación pausada. Presione 'REANUDAR' para continuar.")
    
    
    def cerrar_aplicacion(self):
        """
        Maneja el cierre de la aplicación
        """
        if self.running:
            if messagebox.askyesno("Confirmar", "¿Está seguro de que desea cerrar? Los datos no guardados se perderán."):
                self.running = False
                self.root.destroy()
        else:
            self.root.destroy()
    
    def simular_datos(self):
        """
        Simula datos de balanza
        """
        peso_bruto = round(random.uniform(1.0, 5.0), 3)
        tara = round(random.uniform(0.0, 0.5), 3)
        peso_neto = round(peso_bruto - tara, 3)
        
        timestamp = datetime.now()
        
        datos = {
            'timestamp': timestamp,
            'fecha': timestamp.strftime('%Y-%m-%d'),
            'hora': timestamp.strftime('%H:%M:%S'),
            'lote': self.lote_number,
            'peso_bruto': peso_bruto,
            'tara': tara,
            'peso_neto': peso_neto
        }
        
        self.data_list.append(datos)
        return datos
    
    def guardar_datos_excel(self):
        """
        Guarda los datos en un archivo Excel con formato profesional
        """
        try:
            if not self.data_list:
                print("No hay datos para guardar")
                return False
            
            # Nombre del archivo con número de lote
            nombre_archivo = f"datos_balanza_lote_{self.lote_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Crear estructura de directorios
            directorio_base = "datos_balanza"
            directorio_bines = os.path.join(directorio_base, "Balanza_Bines")
            directorio_cilindros = os.path.join(directorio_base, "Balanza_Cilindros")
            
            # Crear directorios si no existen
            os.makedirs(directorio_bines, exist_ok=True)
            os.makedirs(directorio_cilindros, exist_ok=True)
            
            # Guardar en la carpeta de Bines (IND560)
            ruta_archivo = os.path.join(directorio_bines, nombre_archivo)
            
            # Crear workbook y worksheet
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Balanza"
            
            # Estilos
            header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            data_font = Font(name='Arial', size=11)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Configurar ancho de columnas
            ws.column_dimensions['A'].width = 15  # Fecha
            ws.column_dimensions['B'].width = 12  # Hora
            ws.column_dimensions['C'].width = 15  # Peso Bruto
            ws.column_dimensions['D'].width = 15  # Tara
            ws.column_dimensions['E'].width = 15  # Peso Neto
            ws.column_dimensions['F'].width = 12  # Lote
            
            # Agregar encabezado de empresa
            ws['A1'] = "EMPRESA"
            ws['A1'].font = Font(name='Arial', size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Agregar encabezados de columnas en la fila 3
            headers = ['FECHA', 'HORA', 'PESO BRUTO (kg)', 'TARA (kg)', 'PESO NETO (kg)', 'LOTE']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Agregar datos
            for row, datos in enumerate(self.data_list, 4):
                ws.cell(row=row, column=1, value=datos['fecha']).border = border
                ws.cell(row=row, column=2, value=datos['hora']).border = border
                ws.cell(row=row, column=3, value=datos['peso_bruto']).border = border
                ws.cell(row=row, column=4, value=datos['tara']).border = border
                ws.cell(row=row, column=5, value=datos['peso_neto']).border = border
                ws.cell(row=row, column=6, value=datos['lote']).border = border
                
                # Aplicar formato a las celdas de datos
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar formato numérico a las columnas de peso
            for row in range(4, 4 + len(self.data_list)):
                ws.cell(row=row, column=3).number_format = '0.000'  # Peso Bruto
                ws.cell(row=row, column=4).number_format = '0.000'  # Tara
                ws.cell(row=row, column=5).number_format = '0.000'  # Peso Neto
            
            # Agregar información adicional en la parte inferior
            info_row = 4 + len(self.data_list) + 2
            ws.cell(row=info_row, column=1, value=f"Total de registros: {len(self.data_list)}")
            ws.cell(row=info_row, column=1).font = Font(name='Arial', size=10, italic=True)
            
            ws.cell(row=info_row + 1, column=1, value=f"Archivo generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws.cell(row=info_row + 1, column=1).font = Font(name='Arial', size=10, italic=True)
            
            # Guardar archivo
            wb.save(ruta_archivo)
            
            print(f"✓ Datos guardados en: {ruta_archivo}")
            print(f"✓ Total de registros: {len(self.data_list)}")
            print(f"✓ Formato: Excel (.xlsx) con encabezado de empresa")
            
            return True
            
        except Exception as e:
            print(f"✗ Error al guardar datos: {e}")
            return False
    
    def iniciar_prueba(self):
        """
        Inicia la prueba del sistema
        """
        print("=== SISTEMA DE ADQUISICIÓN DE DATOS DE BALANZA (PRUEBA) ===")
        print("⚠️  MODO PRUEBA: Usando datos simulados")
        print("-" * 50)
        
        # Solicitar número de lote
        self.lote_number = self.solicitar_numero_lote()
        if not self.lote_number:
            print("✗ No se ingresó número de lote. Saliendo...")
            return False
        
        print(f"✓ Número de lote: {self.lote_number}")
        
        # Simular 5 lecturas
        print("\n=== SIMULANDO LECTURAS DE BALANZA ===")
        for i in range(5):
            datos = self.simular_datos()
            print(f"Lectura #{i+1}: Peso={datos['peso_bruto']} kg, Tara={datos['tara']} kg, Neto={datos['peso_neto']} kg")
            time.sleep(1)
        
        # Guardar datos
        print("\n=== GUARDANDO DATOS EN EXCEL ===")
        if self.guardar_datos_excel():
            print("✓ Proceso completado exitosamente")
        else:
            print("✗ Error al guardar datos")
        
        return True

def main():
    """
    Función principal del programa
    """
    # Crear instancia del sistema
    sistema = TestBalanzaSinInput()
    
    # Solicitar número de lote inicial
    sistema.lote_number = sistema.solicitar_numero_lote()
    if not sistema.lote_number:
        print("No se ingresó número de lote. Saliendo...")
        return
    
    # Crear y mostrar interfaz principal
    sistema.crear_interfaz_principal()
    
    # Actualizar información del lote en la interfaz
    if sistema.root:
        # Actualizar el label del lote
        for widget in sistema.root.winfo_children():
            if isinstance(widget, tk.Frame):
                for child in widget.winfo_children():
                    if isinstance(child, tk.Frame):
                        for grandchild in child.winfo_children():
                            if isinstance(grandchild, tk.Label) and "Lote Actual:" in grandchild.cget("text"):
                                grandchild.config(text=f"Lote Actual: {sistema.lote_number}")
                                break
    
    # Iniciar la aplicación
    sistema.root.mainloop()

if __name__ == "__main__":
    main()
