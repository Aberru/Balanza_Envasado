import serial
# import pandas as pd
from datetime import datetime
import os 
import time
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
import threading
import sys

class BalanzaDataAcquisition:
    def __init__(self):
        self.serial_connection = None
        self.data_list = []
        self.running = False
        # Datos de operador
        self.lote_number = ""
        self.codigo = ""
        self.producto = ""
        self.llenado_grueso = ""
        self.llenado_fino = ""
        self.numero_tanque = ""
        # UI
        self.root = None
        self.control_frame = None
        self.status_label = None
        self.serial_status_label = None
        self.operator_info_label = None
        self.data_text = None
        self.contador = 0
        # Buffer de ensamblado por líneas para tramas de 3 líneas
        self.pending_lines = []
        # self.buffer_serial remains for line assembly if needed elsewhere
        self.buffer_serial = ""
        # Remove intermediate pandas-based buffer writing
        # self.contador_excel = 0
        # self.buffer_excel = []
        
    def configurar_puerto_serial(self, puerto="COM1", baudrate=57600, timeout=0.1):
        """
        Configura la conexión serial con la balanza
        """
        try:
            self.serial_connection = serial.Serial(
                port=puerto,
                baudrate=baudrate,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                timeout=timeout
            )
            try:
                self.serial_connection.reset_input_buffer()
            except Exception:
                pass
            if self.serial_status_label:
                self.serial_status_label.config(text="Conexión serial: OK", fg="green")
            print(f"✓ Conexión establecida en {puerto} a {baudrate} baudios, timeout={timeout}s")
            return True
        except Exception as e:
            if self.serial_status_label:
                self.serial_status_label.config(text="Conexión serial: FALLA", fg="red")
            print(f"✗ Error al conectar con la balanza: {e}")
            return False
    
    def leer_datos_balanza(self):
        """
        Lee líneas desde el puerto serial (bloqueante por timeout) y arma tramas de 3 líneas.
        Soporta envíos esporádicos (al presionar 'Imprimir').
        """
        try:
            if self.serial_connection and self.serial_connection.is_open:
                # Leer UNA línea (puede ser vacía si no hay datos durante el timeout)
                linea = self.serial_connection.readline().decode('utf-8', errors='ignore')
                if linea.strip():
                    self.pending_lines.append(linea.strip())
                
                # Si hay al menos 3 líneas acumuladas, procesar una trama
                if len(self.pending_lines) >= 3:
                    frame_lines = self.pending_lines[:3]
                    del self.pending_lines[:3]
                    raw_data = '\n'.join(frame_lines)
                    datos_procesados = self.procesar_datos_balanza(raw_data)
                    if datos_procesados:
                        timestamp = datetime.now()
                        datos_procesados['timestamp'] = timestamp
                        datos_procesados['fecha'] = timestamp.strftime('%Y-%m-%d')
                        datos_procesados['hora'] = timestamp.strftime('%H:%M:%S')
                        # Tomar snapshot de campos operador al momento de la captura
                        datos_procesados['lote'] = self.lote_number
                        datos_procesados['codigo'] = self.codigo
                        datos_procesados['producto'] = self.producto
                        datos_procesados['llenado_grueso'] = self.llenado_grueso
                        datos_procesados['llenado_fino'] = self.llenado_fino
                        datos_procesados['tanque'] = self.numero_tanque
                        self.data_list.append(datos_procesados)
                        return datos_procesados
            return None
        except Exception as e:
            print(f"Error al leer datos: {e}")
            return None
    
    def procesar_datos_balanza(self, raw_data):
        """
        Procesa los datos crudos de la balanza
        Formato esperado:
        0.0 kg 
        0.0 kg T
        0.0 kg N
        """
        try:
            # Dividir por líneas y limpiar
            lineas = raw_data.strip().split('\n')
            lineas_limpias = [linea.strip() for linea in lineas if linea.strip()]
            
            peso_bruto = None
            tara = None
            peso_neto = None
            
            # Procesar cada línea según el formato exacto de HyperTerminal
            for linea in lineas_limpias:
                if 'kg' in linea.lower():
                    try:
                        # Extraer el valor numérico antes de 'kg'
                        partes = linea.split()
                        valor_idx = -1
                        for i, parte in enumerate(partes):
                            if 'kg' in parte.lower():
                                valor_idx = i - 1
                                break
                        
                        if valor_idx >= 0:
                            valor = float(partes[valor_idx])
                            
                            # Identificar el tipo de peso según el formato exacto
                            if 'T' in linea.upper():
                                tara = valor
                            elif 'N' in linea.upper():
                                peso_neto = valor
                            else:
                                # Si no tiene T o N, es peso bruto (primera línea)
                                peso_bruto = valor
                                
                    except (ValueError, IndexError):
                        continue
            
            # Validar que tenemos al menos un valor
            if peso_bruto is None and tara is None and peso_neto is None:
                return None
            
            # Retornar los datos sin corrección
            return {
                'peso_bruto': peso_bruto if peso_bruto is not None else 0.0,
                'tara': tara if tara is not None else 0.0,
                'peso_neto': peso_neto if peso_neto is not None else 0.0,
                'raw_data': raw_data.strip(),  # Datos originales para mostrar
                'timestamp': datetime.now()
            }
            
        except Exception as e:
            print(f"Error al procesar datos: {e}")
            return None
    
    def solicitar_numero_lote(self):
        """
        Solicita el número de lote al usuario mediante interfaz gráfica
        """
        root = tk.Tk()
        root.title("Configuración de Lote")
        root.geometry("400x200")
        root.eval('tk::PlaceWindow . center')
        
        frame = tk.Frame(root, padx=20, pady=20)
        frame.pack(expand=True, fill='both')
        
        tk.Label(frame, text="SISTEMA DE ADQUISICIÓN DE DATOS", 
                font=("Arial", 16, "bold")).pack(pady=10)
        
        tk.Label(frame, text="Ingrese el número de lote:", 
                font=("Arial", 12)).pack(pady=10)
        
        lote_entry = tk.Entry(frame, font=("Arial", 12), width=20)
        lote_entry.pack(pady=10)
        lote_entry.focus()
        
        def on_ok():
            self.lote_number = lote_entry.get()
            if self.lote_number:
                root.destroy()
            else:
                messagebox.showwarning("Advertencia", "Por favor ingrese un número de lote")
        
        def on_cancel():
            self.lote_number = None
            root.destroy()
        
        button_frame = tk.Frame(frame)
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="Aceptar", command=on_ok, 
                font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Cancelar", command=on_cancel, 
                font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)
        
        root.bind('<Return>', lambda e: on_ok())
        root.mainloop()
        return self.lote_number
    
    def crear_interfaz_principal(self):
        """
        Crea la interfaz gráfica principal del sistema
        """
        self.root = tk.Tk()
        self.root.title("Sistema de Adquisición de Datos de Balanza")
        self.root.geometry("900x650")
        self.root.eval('tk::PlaceWindow . center')
        
        # Frame principal
        main_frame = tk.Frame(self.root, padx=10, pady=10)
        main_frame.pack(expand=True, fill='both')
        
        # Título
        title_label = tk.Label(main_frame, text="SISTEMA DE ADQUISICIÓN DE DATOS DE BALANZA", 
                              font=("Arial", 18, "bold"), fg="darkblue")
        title_label.pack(pady=10)
        
        # Información del lote
        info_frame = tk.Frame(main_frame)
        info_frame.pack(fill='x', pady=5)

        # Estado conexión serial
        self.serial_status_label = tk.Label(info_frame, text="Conexión serial: ...", 
                                            font=("Arial", 10), fg="gray")
        self.serial_status_label.pack(side=tk.RIGHT)

        # Información del operador (editable)
        self.operator_info_label = tk.Label(info_frame, text=self._build_operator_info_text(),
                                            font=("Arial", 10), fg="black", justify='left')
        self.operator_info_label.pack(side=tk.LEFT)
        
        # Área de datos
        data_frame = tk.Frame(main_frame)
        data_frame.pack(expand=True, fill='both', pady=10)
        
        tk.Label(data_frame, text="Datos de la Balanza:", 
                font=("Arial", 12, "bold")).pack(anchor='w')
        
        # Text widget con scrollbar
        text_frame = tk.Frame(data_frame)
        text_frame.pack(expand=True, fill='both', pady=5)
        
        self.data_text = tk.Text(text_frame, height=15, font=("Consolas", 10))
        scrollbar = tk.Scrollbar(text_frame, orient="vertical", command=self.data_text.yview)
        self.data_text.configure(yscrollcommand=scrollbar.set)
        
        self.data_text.pack(side=tk.LEFT, expand=True, fill='both')
        scrollbar.pack(side=tk.RIGHT, fill='y')
        
        # Frame de control
        self.control_frame = tk.Frame(main_frame)
        self.control_frame.pack(fill='x', pady=10)
        
        # Botones de control
        button_frame = tk.Frame(self.control_frame)
        button_frame.pack()
        
        self.btn_iniciar = tk.Button(button_frame, text="INICIAR LECTURA", 
                                   command=self.iniciar_lectura, 
                                   font=("Arial", 12, "bold"), 
                                   bg="green", fg="white", width=15)
        self.btn_iniciar.pack(side=tk.LEFT, padx=5)
        
        # Cambiar texto a DETENER Y GUARDAR
        self.btn_detener = tk.Button(button_frame, text="DETENER Y GUARDAR", 
                                   command=self.detener_y_guardar, 
                                   font=("Arial", 12, "bold"), 
                                   bg="red", fg="white", width=20, state='disabled')
        self.btn_detener.pack(side=tk.LEFT, padx=5)
        
        # Botón editar datos
        self.btn_editar = tk.Button(button_frame, text="Editar datos", 
                                   command=self.editar_datos_dialog, 
                                   font=("Arial", 12, "bold"), 
                                   bg="blue", fg="white", width=15)
        self.btn_editar.pack(side=tk.LEFT, padx=5)

        # Botón N° Tanque (entrada rápida)
        self.btn_tanque = tk.Button(button_frame, text="N° Tanque", 
                                   command=self.cambiar_tanque_dialog, 
                                        font=("Arial", 12, "bold"), 
                                   bg="orange", fg="white", width=12)
        self.btn_tanque.pack(side=tk.LEFT, padx=5)
        
        
        # Status bar
        self.status_label = tk.Label(main_frame, text="Sistema listo.", 
                                   font=("Arial", 10), fg="blue", relief=tk.SUNKEN, anchor='w')
        self.status_label.pack(fill='x', pady=5)
        
        # Configurar cierre de ventana
        self.root.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion)
    
    def _build_operator_info_text(self):
        return (f"Lote: {self.lote_number or '-'}    Código: {self.codigo or '-'}    Producto: {self.producto or '-'}\n"
                f"Llenado grueso (kg): {self.llenado_grueso or '-'}    Llenado fino (kg): {self.llenado_fino or '-'}    N° Tanque: {self.numero_tanque or '-'}")

    def actualizar_operator_info(self):
        if self.operator_info_label:
            self.operator_info_label.config(text=self._build_operator_info_text())
            self.root.update()

    def editar_datos_dialog(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Editar datos")
        dlg.transient(self.root)
        dlg.grab_set()
        frm = tk.Frame(dlg, padx=15, pady=15)
        frm.pack()
        entries = {}
        campos = [
            ("Lote", self.lote_number),
            ("Código", self.codigo),
            ("Producto", self.producto),
            ("Llenado grueso (kg)", self.llenado_grueso),
            ("Llenado fino (kg)", self.llenado_fino),
        ]
        for idx, (label, val) in enumerate(campos):
            tk.Label(frm, text=label+":", font=("Arial", 10)).grid(row=idx, column=0, sticky='e', pady=3, padx=5)
            ent = tk.Entry(frm, font=("Arial", 10), width=28)
            ent.grid(row=idx, column=1, pady=3, padx=5)
            ent.insert(0, val or "")
            entries[label] = ent
        def on_ok():
            self.lote_number = entries["Lote"].get().strip()
            self.codigo = entries["Código"].get().strip()
            self.producto = entries["Producto"].get().strip()
            self.llenado_grueso = entries["Llenado grueso (kg)"].get().strip()
            self.llenado_fino = entries["Llenado fino (kg)"].get().strip()
            self.actualizar_operator_info()
            dlg.destroy()
        tk.Button(frm, text="Aceptar", command=on_ok, bg="green", fg="white", width=12).grid(row=len(campos), column=0, pady=10)
        tk.Button(frm, text="Cancelar", command=dlg.destroy, bg="gray", fg="white", width=12).grid(row=len(campos), column=1, pady=10)

    def cambiar_tanque_dialog(self):
        valor = simpledialog.askstring("N° Tanque", "Ingrese N° Tanque:", parent=self.root)
        if valor is not None:
            self.numero_tanque = valor.strip()
            self.actualizar_operator_info()
    
    
    def actualizar_status(self, mensaje):
        """
        Actualiza el mensaje de estado
        """
        if self.status_label:
            self.status_label.config(text=mensaje)
            self.root.update()
    
    def agregar_dato(self, datos):
        """
        Agrega un dato al área de texto mostrando los datos exactos como en HyperTerminal
        """
        if self.data_text:
            self.contador += 1
            timestamp = datetime.now().strftime('%H:%M:%S')
            # Mostrar los datos exactos como llegan de la balanza
            texto = f"[{timestamp}] Lectura #{self.contador}:\n{datos['raw_data']}\n"
            self.data_text.insert(tk.END, texto)
            self.data_text.see(tk.END)
            self.root.update()
    

    def iniciar_lectura(self):
        """
        Inicia la lectura de datos en un hilo separado
        """
        if not self.serial_connection or not self.serial_connection.is_open:
            if not self.configurar_puerto_serial():
                messagebox.showerror("Error", "No se pudo conectar con la balanza")
                return
        
        # Limpiar buffers para no perder la primera impresión
        self.pending_lines = []
        try:
            self.serial_connection.reset_input_buffer()
        except Exception:
            pass
        
        self.running = True
        self.btn_iniciar.config(state='disabled')
        self.btn_detener.config(state='normal')
        # Iniciar hilo de lectura
        self.thread_lectura = threading.Thread(target=self.hilo_lectura, daemon=True)
        self.thread_lectura.start()
        self.actualizar_status("Leyendo datos de la balanza... Presione 'DETENER Y GUARDAR' para generar Excel.")
    
    def hilo_lectura(self):
        """
        Hilo principal de lectura de datos optimizado para respuesta en tiempo real
        """
        while self.running:
            try:
                datos = self.leer_datos_balanza()
                if datos:
                    self.agregar_dato(datos)
                # Ritmo de lectura fijo
                time.sleep(0.1)
            except Exception as e:
                self.actualizar_status(f"Error en lectura: {e}")
                break
    
    def detener_y_guardar(self):
        """
        Detiene la lectura y guarda los datos
        """
        self.running = False
        self.actualizar_status("Deteniendo lectura y guardando...")
        
        # Guardar datos si hay
        if self.data_list:
            if self.guardar_datos_excel():
                self.actualizar_status("Datos guardados exitosamente.")
            else:
                self.actualizar_status("Error al guardar datos.")
        else:
            self.actualizar_status("No hay datos para guardar.")
        
        # Mantener app abierta
        self.btn_iniciar.config(state='normal')
        self.btn_detener.config(state='disabled')
    
    def pausar_adquisicion(self):
        """
        Pausa la adquisición de datos
        """
        if self.running:
            # Pausar la adquisición
            self.running = False
            self.actualizar_status("Adquisición pausada. Presione 'INICIAR LECTURA' para continuar.")
            self.btn_iniciar.config(state='normal')
            self.btn_pausar.config(state='disabled')
            self.btn_detener.config(state='normal')
        else:
            # Reanudar la adquisición
            self.running = True
            self.actualizar_status("Reanudando adquisición...")
            self.btn_iniciar.config(state='disabled')
            self.btn_pausar.config(state='normal')
            self.btn_detener.config(state='normal')
            
            # Reiniciar hilo de lectura
            self.thread_lectura = threading.Thread(target=self.hilo_lectura, daemon=True)
            self.thread_lectura.start()
    
    def actualizar_label_lote(self):
        """
        Actualiza el label del lote en la interfaz
        """
        if self.root:
            # Buscar y actualizar el label del lote
            for widget in self.root.winfo_children():
                if isinstance(widget, tk.Frame):
                    for child in widget.winfo_children():
                        if isinstance(child, tk.Frame):
                            for grandchild in child.winfo_children():
                                if isinstance(grandchild, tk.Label) and "Lote Actual:" in grandchild.cget("text"):
                                    grandchild.config(text=f"Lote Actual: {self.lote_number}")
                                    return
    
    def cerrar_aplicacion(self):
        """
        Maneja el cierre de la aplicación
        """
        if self.running:
            if messagebox.askyesno("Confirmar", "¿Está seguro de que desea cerrar? Los datos no guardados se perderán."):
                self.running = False
                if self.serial_connection and self.serial_connection.is_open:
                    self.serial_connection.close()
                self.root.destroy()
        else:
            if self.serial_connection and self.serial_connection.is_open:
                self.serial_connection.close()
            self.root.destroy()
    
    def guardar_datos_excel(self):
        """
        Guarda los datos en un archivo Excel con formato profesional
        - Encabezado superior ordenado con datos del operador y fecha/hora
        - Tabla de datos con columnas: FECHA, HORA, PESO BRUTO, TARA, PESO NETO, N° TANQUE
        """
        try:
            if not self.data_list:
                print("No hay datos para guardar")
                return False
            
            nombre_archivo = f"datos_balanza_lote_{self.lote_number or 'NA'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            directorio_base = "datos_balanza"
            directorio_bines = os.path.join(directorio_base, "Balanza_Bines")
            directorio_cilindros = os.path.join(directorio_base, "Balanza_Cilindros")
            os.makedirs(directorio_bines, exist_ok=True)
            os.makedirs(directorio_cilindros, exist_ok=True)
            ruta_archivo = os.path.join(directorio_bines, nombre_archivo)
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Balanza"
            
            # Estilos
            title_font = Font(name='Arial', size=16, bold=True)
            label_font = Font(name='Arial', size=11, bold=True)
            value_font = Font(name='Arial', size=11)
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Encabezado profesional
            ws.merge_cells('A1:F1')
            ws['A1'] = 'REGISTRO DE PESOS - BALANZA'
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:F2')
            ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].alignment = Alignment(horizontal='center')

            # Bloque de información ingresada por el operador (ordenado)
            info_rows = [
                ("Lote", self.lote_number or ''),
                ("Código", self.codigo or ''),
                ("Producto", self.producto or ''),
                ("Llenado grueso (kg)", self.llenado_grueso or ''),
                ("Llenado fino (kg)", self.llenado_fino or ''),
                ("N° Tanque", self.numero_tanque or ''),
            ]
            start_row = 4
            for idx, (label, value) in enumerate(info_rows):
                row = start_row + idx
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = label_font
                ws[f'A{row}'].alignment = Alignment(horizontal='left')

                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = value_font
                ws[f'B{row}'].alignment = Alignment(horizontal='left')

            # Fila separadora
            table_header_row = start_row + len(info_rows) + 1

            # Encabezados de la tabla
            headers = ['FECHA', 'HORA', 'PESO BRUTO (kg)', 'TARA (kg)', 'PESO NETO (kg)', 'N° TANQUE']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=table_header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin

            # Datos
            first_data_row = table_header_row + 1
            for row_offset, datos in enumerate(self.data_list):
                r = first_data_row + row_offset
                ws.cell(row=r, column=1, value=datos.get('fecha')).border = border_thin
                ws.cell(row=r, column=2, value=datos.get('hora')).border = border_thin
                ws.cell(row=r, column=3, value=datos.get('peso_bruto')).border = border_thin
                ws.cell(row=r, column=4, value=datos.get('tara')).border = border_thin
                ws.cell(row=r, column=5, value=datos.get('peso_neto')).border = border_thin
                ws.cell(row=r, column=6, value=datos.get('tanque')).border = border_thin
                for c in range(1, 7):
                    cell = ws.cell(row=r, column=c)
                    cell.font = value_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Anchos de columna
            widths = [15, 12, 16, 14, 16, 14]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # Formatos numéricos
            for r in range(first_data_row, first_data_row + len(self.data_list)):
                ws.cell(row=r, column=3).number_format = '0.000'
                ws.cell(row=r, column=4).number_format = '0.000'
                ws.cell(row=r, column=5).number_format = '0.000'


            wb.save(ruta_archivo)
            print(f"✓ Datos guardados en: {ruta_archivo}")
            return True
        except Exception as e:
            print(f"✗ Error al guardar datos: {e}")
            return False
    
    def iniciar_adquisicion(self, puerto="COM1", baudrate=57600, intervalo=0.1):
        """
        Inicia el proceso de adquisición de datos
        """
        print("=== SISTEMA DE ADQUISICIÓN DE DATOS DE BALANZA ===")
        print(f"Puerto: {puerto}")
        print(f"Baudrate: {baudrate}")
        print(f"Intervalo de lectura: {intervalo} segundos")
        print("-" * 50)
        
        # Solicitar número de lote
        self.lote_number = self.solicitar_numero_lote()
        if not self.lote_number:
            print("✗ No se ingresó número de lote. Saliendo...")
            return False
        
        print(f"✓ Número de lote: {self.lote_number}")
        
        # Configurar puerto serial
        if not self.configurar_puerto_serial(puerto, baudrate):
            return False
        
        # Iniciar adquisición
        self.running = True
        contador = 0
        
        try:
            print("\n=== INICIANDO ADQUISICIÓN DE DATOS ===")
            print("Presione Ctrl+C para detener y guardar datos...")
            
            while self.running:
                datos = self.leer_datos_balanza()
                if datos:
                    contador += 1
                    print(f"Lectura #{contador}: Peso={datos['peso_bruto']} kg, Tara={datos['tara']} kg, Neto={datos['peso_neto']} kg")
                
                time.sleep(intervalo)
                
        except KeyboardInterrupt:
            print("\n\n=== DETENIENDO ADQUISICIÓN ===")
            self.running = False
        
        finally:
            # Cerrar conexión serial
            if self.serial_connection and self.serial_connection.is_open:
                self.serial_connection.close()
                print("✓ Conexión serial cerrada")
            
            # Guardar datos
            if self.data_list:
                print("\n=== GUARDANDO DATOS ===")
                if self.guardar_datos_excel():
                    print("✓ Proceso completado exitosamente")
                else:
                    print("✗ Error al guardar datos")
            else:
                print("✗ No se capturaron datos")
        
        return True

def main():
    """
    Función principal del programa
    """
    sistema = BalanzaDataAcquisition()
    # Crear y mostrar interfaz principal (sin solicitar lote)
    sistema.crear_interfaz_principal()
    sistema.root.mainloop()

if __name__ == "__main__":
    main()