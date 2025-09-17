import tkinter as tk
from tkinter import messagebox, simpledialog
from datetime import datetime
import os
import time
import random

class TestBalanzaExcel:
    def __init__(self):
        self.data_list = []
        self.lote_number = ""
        
    def solicitar_numero_lote(self):
        """
        Solicita el número de lote al usuario mediante interfaz gráfica
        """
        root = tk.Tk()
        root.title("Sistema de Balanza - Prueba Excel")
        root.geometry("400x300")
        
        # Centrar ventana
        root.eval('tk::PlaceWindow . center')
        
        # Crear interfaz
        frame = tk.Frame(root, padx=20, pady=20)
        frame.pack(expand=True, fill='both')
        
        tk.Label(frame, text="SISTEMA DE ADQUISICIÓN DE DATOS", 
                font=("Arial", 16, "bold")).pack(pady=10)
        
        tk.Label(frame, text="Ingrese el número de lote:", 
                font=("Arial", 12)).pack(pady=10)
        
        lote_entry = tk.Entry(frame, font=("Arial", 12), width=20)
        lote_entry.pack(pady=10)
        lote_entry.focus()
        
        def on_ok():
            self.lote_number = lote_entry.get()
            if self.lote_number:
                root.destroy()
            else:
                messagebox.showwarning("Advertencia", "Por favor ingrese un número de lote")
        
        def on_cancel():
            self.lote_number = None
            root.destroy()
        
        button_frame = tk.Frame(frame)
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="Aceptar", command=on_ok, 
                font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Cancelar", command=on_cancel, 
                font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)
        
        # Enter para aceptar
        root.bind('<Return>', lambda e: on_ok())
        
        root.mainloop()
        return self.lote_number
    
    def simular_datos(self):
        """
        Simula datos de balanza
        """
        peso_bruto = round(random.uniform(1.0, 5.0), 3)
        tara = round(random.uniform(0.0, 0.5), 3)
        peso_neto = round(peso_bruto - tara, 3)
        
        timestamp = datetime.now()
        
        datos = {
            'timestamp': timestamp,
            'fecha': timestamp.strftime('%Y-%m-%d'),
            'hora': timestamp.strftime('%H:%M:%S'),
            'lote': self.lote_number,
            'peso_bruto': peso_bruto,
            'tara': tara,
            'peso_neto': peso_neto
        }
        
        self.data_list.append(datos)
        return datos
    
    def guardar_datos_excel(self):
        """
        Guarda los datos en un archivo Excel con formato personalizado
        """
        try:
            if not self.data_list:
                print("No hay datos para guardar")
                return False
            
            # Nombre del archivo con número de lote
            nombre_archivo = f"datos_balanza_lote_{self.lote_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Crear directorio si no existe
            directorio = "datos_balanza"
            if not os.path.exists(directorio):
                os.makedirs(directorio)
            
            ruta_archivo = os.path.join(directorio, nombre_archivo)
            
            # Crear workbook y worksheet
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Balanza"
            
            # Estilos
            header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            data_font = Font(name='Arial', size=11)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Configurar ancho de columnas
            ws.column_dimensions['A'].width = 15  # Fecha
            ws.column_dimensions['B'].width = 12  # Hora
            ws.column_dimensions['C'].width = 15  # Peso Bruto
            ws.column_dimensions['D'].width = 15  # Tara
            ws.column_dimensions['E'].width = 15  # Peso Neto
            ws.column_dimensions['F'].width = 12  # Lote
            
            # Agregar encabezado de empresa
            ws['A1'] = "EMPRESA"
            ws['A1'].font = Font(name='Arial', size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Agregar encabezados de columnas en la fila 3
            headers = ['FECHA', 'HORA', 'PESO BRUTO (kg)', 'TARA (kg)', 'PESO NETO (kg)', 'LOTE']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Agregar datos
            for row, datos in enumerate(self.data_list, 4):
                ws.cell(row=row, column=1, value=datos['fecha']).border = border
                ws.cell(row=row, column=2, value=datos['hora']).border = border
                ws.cell(row=row, column=3, value=datos['peso_bruto']).border = border
                ws.cell(row=row, column=4, value=datos['tara']).border = border
                ws.cell(row=row, column=5, value=datos['peso_neto']).border = border
                ws.cell(row=row, column=6, value=datos['lote']).border = border
                
                # Aplicar formato a las celdas de datos
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar formato numérico a las columnas de peso
            for row in range(4, 4 + len(self.data_list)):
                ws.cell(row=row, column=3).number_format = '0.000'  # Peso Bruto
                ws.cell(row=row, column=4).number_format = '0.000'  # Tara
                ws.cell(row=row, column=5).number_format = '0.000'  # Peso Neto
            
            # Agregar información adicional en la parte inferior
            info_row = 4 + len(self.data_list) + 2
            ws.cell(row=info_row, column=1, value=f"Total de registros: {len(self.data_list)}")
            ws.cell(row=info_row, column=1).font = Font(name='Arial', size=10, italic=True)
            
            ws.cell(row=info_row + 1, column=1, value=f"Archivo generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws.cell(row=info_row + 1, column=1).font = Font(name='Arial', size=10, italic=True)
            
            # Guardar archivo
            wb.save(ruta_archivo)
            
            print(f"✓ Datos guardados en: {ruta_archivo}")
            print(f"✓ Total de registros: {len(self.data_list)}")
            print(f"✓ Formato: Excel (.xlsx) con encabezado de empresa")
            
            return True
            
        except Exception as e:
            print(f"✗ Error al guardar datos: {e}")
            return False
    
    def iniciar_prueba(self):
        """
        Inicia la prueba del sistema
        """
        print("=== SISTEMA DE ADQUISICIÓN DE DATOS DE BALANZA (PRUEBA EXCEL) ===")
        print("⚠️  MODO PRUEBA: Usando datos simulados")
        print("-" * 50)
        
        # Solicitar número de lote
        self.lote_number = self.solicitar_numero_lote()
        if not self.lote_number:
            print("✗ No se ingresó número de lote. Saliendo...")
            return False
        
        print(f"✓ Número de lote: {self.lote_number}")
        
        # Simular 5 lecturas
        print("\n=== SIMULANDO LECTURAS DE BALANZA ===")
        for i in range(5):
            datos = self.simular_datos()
            print(f"Lectura #{i+1}: Peso={datos['peso_bruto']} kg, Tara={datos['tara']} kg, Neto={datos['peso_neto']} kg")
            time.sleep(1)
        
        # Guardar datos
        print("\n=== GUARDANDO DATOS EN EXCEL ===")
        if self.guardar_datos_excel():
            print("✓ Proceso completado exitosamente")
        else:
            print("✗ Error al guardar datos")
        
        return True

def main():
    """
    Función principal del programa
    """
    sistema = TestBalanzaExcel()
    sistema.iniciar_prueba()
    
    # Pausa antes de cerrar
    print("\nPresione Enter para cerrar...")
    try:
        input()
    except:
        print("Programa finalizado.")

if __name__ == "__main__":
    main()
